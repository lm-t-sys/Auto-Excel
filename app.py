#!user/bin/python
#coding : utf-8
from flask import Flask,render_template,request,redirect,url_for
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from IPython.display import HTML
from openpyxl import load_workbook
from werkzeug.utils import secure_filename 
import os
UPLOAD_FOLDER="static/excelfolder/"
app=Flask(__name__)
app.config['UPLOAD_FOLDER']=UPLOAD_FOLDER
@app.route('/create',methods=['GET','POST'])
def create():
    if request.method=="GET":
        return render_template('create.html')
    else:    
        a=request.form.get("indexs")
        b=request.form.get("columns")
        c=request.form.get("values")
        btn1=request.form.get("btn1")
        btn2=request.form.get("btn2")    
        on=False
        btn1=btn1.strip()
        btn2=btn2.strip()
        on=False

        if len(btn2)>=1:
            on=True
        a3,b3=[],[]
        a2=a.split(',')
        b2=b.split(",")
        for aa in a2:
            if a=="":
                break
            a3.append(aa)

        for bb in b2:
            if b=="":
                break
            b3.append(bb) 
        f=[] 

        d=c.strip().split(",") 
        for i in d:
            f.append(i)    
        f=np.array(f) 
        len(b3)
        a3_size=len(a3)
        b3_size=len(b3)
        if on==True:
            f=np.array(f).reshape(int(f.size/int(btn2)),int(btn2)) 
        elif b3_size>1 and on==False:
            f=np.array(f).reshape(int(f.size/b3_size),b3_size) 
        else:
            f=np.array(f) 
        a3_size  
        if a3_size<1 and b3_size<1:
            c2=pd.DataFrame(f,columns=None,index=None)
        elif b3_size>0 and a3_size<1:
            c2=pd.DataFrame(f,columns=b3,index=None)
        else:
            c2=pd.DataFrame(f,columns=b3,index=a3)  
    

        if a3_size<1 and b3_size<1:
            c2.to_excel('static/excelfolder/to_csv.xlsx',index=False,header=False)
        if a3_size<1 and b3_size>0:
            c2.to_excel('static/excelfolder/to_csv.xlsx',index=False)
        if a3_size>0 and b3_size>0:
            c2.to_excel('static/excelfolder/to_csv.xlsx')  
        ANS=HTML(c2.to_html())      

        return render_template('create.html',posts=ANS)

@app.route("/add",methods=["GET","POST"])
def add():
    if request.method=="GET":
        return render_template("add.html")
    else:
        file=request.files["file"]
        name=request.form.get("sheet_name")
        col=request.form.get('col')
        value=request.form.get('value')

        filename=secure_filename(file.filename)
        file.save(os.path.join(app.config['UPLOAD_FOLDER'],filename))
        files="static/excelfolder/to_csv.xlsx"
        wb=load_workbook(filename=files)
        sheet=wb[name]
        sheet["{}{}".format(col,sheet.max_row+1)]=value
        wb.save(files)
        wb.close()
        
        wb=pd.read_excel(files,header=0)
        a2=wb.to_html()
        return render_template("add.html",data=a2) 

@app.route('/auto',methods=['GET','POST'])  
def auto():
    if request.method=='GET':
        return render_template('auto.html')
    else:
        filee=request.files['filee']
        filenam=secure_filename(filee.filename)
        filee.save(os.path.join(app.config['UPLOAD_FOLDER'],filenam))
        an="static/excelfolder/to_auto.xlsx"
        text1=request.form.get('change')
        text2=request.form.get('inputs')
        
        if "\n" in text2:
            text2a=text2.split("\n")
        else:
            text2a=text2
            
        text1a=text1.split(',')

        wb=load_workbook(filename=an)
        sheetnames=wb.sheetnames
        for i,sheet_name in enumerate(sheetnames):
            moji=text2a[i].split(',')
            sheet=wb[sheet_name]
            for j in range(len(text1a)):
                sheet[text1a[j]]=moji[j]
        wb.save(an)  
        wb.close()
        wb=pd.read_excel(filee,header=0)
        aa3=wb.to_html()
        return render_template('auto.html',data=aa3)             
        
         


@app.route('/',methods=['GET','POST'])
def index():
    if request.method=='GET':

        return render_template("index.html")
    else:
         if request.form.get("hide")=="新規作成":
            posts=request.form.get("hide")
            return render_template('create.html',posts=posts)
         else:
            return redirect('/')      
      

if __name__=="__main__":
    app.run(debug=True)    